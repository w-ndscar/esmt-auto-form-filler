#!/usr/bin/env python

import time
import datetime
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from pynput import keyboard
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox

# This application is used to automate the process of filling some web forms with data from an Excel sheet.

def only_integers(char):
    """Function to validate that input is an integer or empty."""
    return char.isdigit() or char == ""

# Select the functionality from the drop down
def select_functionality(choice):
    def choicesoneandtwo():
        app.geometry("650x550")
        app.title(f"ESMT - Automation - {choice}")
        frame_3.pack_forget()
        frame_1.pack(side="top", pady=20, padx=20, ipady=20, ipadx=20, fill="both", expand=True, anchor="center")
        frame_2.pack(side="bottom", pady=20, padx=20, ipady=20, ipadx=20, fill="both", expand=True, anchor="center")
    
    choice = function_box.get()
    if choice == "Add Drawings":
        CTkMessagebox(title="Info", message="'Add Drawings' is selected", icon="info")
        choicesoneandtwo()
    elif choice == "Active to Completed":
        CTkMessagebox(title="Info", message="'Active to Completed' is selected", icon="info")
        choicesoneandtwo()
    elif choice == "Convert to Excel (WorkTime)":
        CTkMessagebox(title="Info", message="'Convert to Excel (WorkTime)' is selected", icon="info")
        app.title(f"ESMT - Automation - {choice}")
        # Hide the frames for this functionality
        frame_2.pack_forget()
        frame_1.pack(side="top", pady=20, padx=20, ipady=20, ipadx=20, fill="both", expand=True, anchor="center")
        frame_3.pack(side="bottom", pady=20, padx=20, ipady=20, ipadx=20, fill="both", expand=True, anchor="center")
        app.geometry("650x450")
    return choice

# Launching Chrome
def launch_chrome():
    global driver
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--start-maximized")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    driver.get("https://projectmanagement-8d5c4.web.app/")
    #driver.maximize_window()
    CTkMessagebox(title="Info", message="ESMT Launched. Please Log in -> Select the Project -> Click on Drawings Tab. Check Instructions for more info")
    time.sleep(2)  # Wait a bit for Chrome to start

# Browse file logic
def browse_file(target_var):
    filename = ctk.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.csv")])
    target_var.set(filename)

# Hotkey Wait
def wait_for_key():
    # Function to wait for the 'Esc' key to be pressed.
    def on_press(key):
        if key == keyboard.Key.esc:
            return False  # Stop listener

    # Collect events until released
    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()

# Instructions
def show_instructions():
    instructions_window = ctk.CTkToplevel(app)
    instructions_window.title("Instructions")
    instructions_window.geometry("740x600")
    instructions_text = (
        "ESMT - Auto Form Filler v1.0.5-beta\n\n"
        "Pre-run check: Select the correct project and ensure all elements are present.\n"
        "If any are missing, create them before running the automation — otherwise, the 'Elements' section will appear empty.\n\n"
        "1. Select the desired function from the dropdown\n"
        "2. Click 'Launch ESMT' and log in to the website\n"
        "3. Select the project and go to the Drawings tab\n"
        "4. Use 'Browse' to select your Excel file\n"
        "5. Enter the sheet name, starting and ending row number\n"
        "6. Click 'Start Automation' to begin\n"
        "7. For 'Convert to Excel', just select the CSV File and click Convert"
        "\nTip: After entering the values, it'll wait \n"
        "Press 'Esc' key after each entry to proceed to the next \n"
        "DO NOT click Submit/New Buttons. It will be done automatically\n"
        "For 'Active to Completed', manually click on 'Yes, update it!' button and then press Esc\n"
        "\n\n\n\n"
        "Created with curiosity by Arun"
    )
    #ctk.CTkLabel(instructions_window, text="How to Use", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10))
    ctk.CTkTextbox(instructions_window, width=720, height=570, font=ctk.CTkFont(size=16), border_color="#d3eef7").pack(padx=20, pady=5)
    textbox = instructions_window.winfo_children()[-1]
    textbox.insert("1.0", instructions_text)
    textbox.configure(state="disabled")

# Display missing drawings - Active to Completed
def show_missing_drawings(missing_drawings):
    instructions_window = ctk.CTkToplevel(app)
    instructions_window.title("Missing Drawings")
    instructions_window.geometry("740x600")
    instructions_text = "Note: Missing Drawings are also exported to the current folder\n\nMissing Drawing Numbers: \n\n" + "\n".join(missing_drawings)
    ctk.CTkTextbox(instructions_window, width=720, height=570, font=ctk.CTkFont(size=16), border_color="#d3eef7").pack(padx=20, pady=5)
    textbox = instructions_window.winfo_children()[-1]
    textbox.insert("1.0", instructions_text)
    textbox.configure(state="disabled")

# Loads the excel
def load_excel(file_path, sheet_name, start_row, end_row):
    try:
        print("Load excel function called")
        start_row = int(start_row)
        end_row = int(end_row)
        end_row += 1  # Adjusted for 1-based indexing in Excel

        col1 = []
        col2 = []
        col3 = []
        col4 = []
        col5 = []
        col6 = []
        
        # Read data from the specified range
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        print("Reached the Excel file reading part")
        def read_column(sheet, col, start_row, end_row, date_format=None):
            values = []
            for row in range(start_row, end_row):
                char = get_column_letter(col)
                cell_name = char + str(row)
                cell_value = sheet[cell_name].value
                if date_format and cell_value:
                    try:
                        cell_value = cell_value.strftime(date_format)
                    except AttributeError:
                        pass # Not a date, keep as is
                values.append(cell_value)
            return values

        col1 = read_column(sheet, 2, start_row, end_row)
        col2 = read_column(sheet, 3, start_row, end_row)
        col3 = read_column(sheet, 4, start_row, end_row, date_format='%d-%m-%Y')
        col4 = read_column(sheet, 5, start_row, end_row)
        col5 = read_column(sheet, 6, start_row, end_row)
        col6 = read_column(sheet, 7, start_row, end_row)
        print("Data read from Excel file")

        data = {
            'col1': col1,
            'col2': col2,
            'col3': col3,
            'col4': col4,
            'col5': col5,
            'col6': col6
        }

        return data

    except Exception as e:
        CTkMessagebox(title="Error", message=f"Failed to read Excel file: {e}", icon="cancel")
        return None

# Converts CSV to XLSX and reformats (custom)
def convert_excel(file_path):
    
    def timedelta_to_excel_time(td):
        total_seconds = td.total_seconds()
        # Excel time = fraction of a day
        return total_seconds / 86400  # 86400 seconds in a day

    # Store and reformat
    print(f"File Path: '{file_path}'")
    df = pd.read_csv(file_path, encoding='utf-8', header=2, skiprows=[0, 1], sep=r'[\t,]', engine='python')
    print("Reached the CSV file reading part")
    print(df.columns)
    
    # Drop Columns
    # cols_to_drop = [2,3,4,5,6,9,10]
    # df = df.drop(df.columns[cols_to_drop], axis=1)

    # Defining column indexes to keep (0-based indexing)
    df = df.iloc[:, [0,1,7,8]]  # Keep only the columns we need
    df.columns = ['Employee', 'Date', 'Active time', 'Idle time']

    print("Reached the date conversion part")
    # Convert and format date column
    df['Date'] = pd.to_datetime(df['Date'], format="%d/%m/%Y", errors='coerce')
    
    print("Reached the time conversion part")
    # Convert and format time columns
    df['Active time'] = pd.to_timedelta(df['Active time'])
    df['Idle time'] = pd.to_timedelta(df['Idle time'])
    
    df['Active time'] = df['Active time'].apply(timedelta_to_excel_time)
    df['Idle time'] = df['Idle time'].apply(timedelta_to_excel_time)

    # Write to a new Excel file
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    df.to_excel(f"worktime_converted_{timestamp}.xlsx", index=False)

    # Excel Manipulation
    workbook = openpyxl.load_workbook(f"worktime_converted_{timestamp}.xlsx")
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    date_col = headers.index("Date") + 1
    active_time_col = headers.index("Active time") + 1
    idle_time_col = headers.index("Idle time") + 1

    # Applying date and time format
    for row in sheet.iter_rows(min_row=2):
        row[date_col - 1].number_format = "dd/mmm/yyyy"
        row[active_time_col - 1].number_format = "h:mm:ss"
        row[idle_time_col - 1].number_format = "h:mm:ss"

    workbook.save(f"worktime_converted_{timestamp}.xlsx")

    CTkMessagebox(title="Success", message="Conversion completed successfully! Saved to the current folder", icon="check")

# Add drawing function
def start_adding_dwg(data):

    global driver
    
    wait = WebDriverWait(driver, 10)

    element = data['col1']
    sh_size = data['col2']
    st_date = data['col3']
    dwg_name = data['col4']
    dwg_desc = data['col5']
    revision = data['col6']
    print("Starting to add drawings with the following data:")

    def fill_input(by, selector, value, clear=True):
        elem = wait.until(EC.presence_of_element_located((by, selector)))
        if clear:
            elem.clear()
        elem.send_keys(value)

    def select_ng_autocomplete(input_xpath, value, wait):
        # 1. Find the input and type the value
        input_elem = wait.until(EC.element_to_be_clickable((By.XPATH, input_xpath)))
        input_elem.clear()
        input_elem.send_keys(value)
        sleep(1)  # Wait for dropdown to populate (adjust as needed)

        # 2. Press DOWN and ENTER to select the first matching option
        input_elem.send_keys(Keys.ARROW_DOWN)
        input_elem.send_keys(Keys.ENTER)
        sleep(0.5)

    for i in range(len(element)):
        #Click - New Button
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div/div[2]/div/button[1]')))
        elem2.click()
        sleep(1)
        
        print("\n")
        print(" | ")

        #Fields

        #Element
        print (element[i], " | ")
        sleep(0.5)
        # fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[1]/ng-autocomplete/div[1]/div[1]/input', element[i])
        select_ng_autocomplete('/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[1]/ng-autocomplete/div/div[1]/input', element[i], wait)

        #Sheet Size - Dropdown
        print (sh_size[i], " | ")
        sleep(0.5)
        Select(driver.find_element(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[2]/select')).select_by_visible_text(sh_size[i])
        sleep(0.5)

        #Scheduled Date
        print (st_date[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[3]/input', st_date[i])

        #Drawing name/number
        print (dwg_name[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[4]/input', dwg_name[i])

        #Drawing Description
        print (dwg_desc[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[5]/input', dwg_desc[i])
        sleep(0.5)
        
        #Revision
        print (revision[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/form/div[6]/input', revision[i])

        #Hotkey to wait before proceeding to Submit
        wait_for_key()

        #Click - Submit
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/div/button[2]')))
        elem2.click()
        sleep(0.5)

# Active to completed function
def start_active_to_completed(data):

    global driver
    
    wait = WebDriverWait(driver, 10)

    dwg_no = data['col1']
    dwg_date = data['col3']
    missing_dwg = []

    print("Starting active to completed with the following data:")

    def fill_input(by, selector, value, clear=True):
        elem = wait.until(EC.presence_of_element_located((by, selector)))
        if clear:
            elem.clear()
        elem.send_keys(value)

    for i in range(len(dwg_no)):
        # Fill Search Input
        print(dwg_no[i], " | ")
        sleep(1.5)
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div/div[3]/input', dwg_no[i])

        # Click Search Button
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div/div[3]/button[1]')))
        elem2.click()

        # Click 'Set Drawings as Completed' Button
        elems = driver.find_elements(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div/div[5]/div[11]/button[6]')
        if elems:
            elem2 = elems[0]
            elem2.click()
        else:
            missing_dwg.append(dwg_no[i])
            continue  # Skip to the next drawing if not found

        #Scheduled Date
        print (dwg_date[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[2]/input', dwg_date[i])

        #Click - Save Button
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div[3]/button[1]')))
        elem2.click()
        sleep(0.5)

        #Hotkey to wait before proceeding to Confirm Update
        wait_for_key()
        sleep(1)

    # Display missing drawings if any
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    f = open(f"missing_drawings_{timestamp}.txt", "w")
    f.write(f"Last Run: {timestamp} \n\nMissing drawings:\n" + "\n".join(missing_dwg))
    f.close()
    show_missing_drawings(missing_dwg)

# Start Automation function
def start_button_clicked():
    try:
        # Ensure the functionality is selected then proceed with the relevant automation
        if function_box.get() == "Add Drawings":
            start_adding_dwg(load_excel(file_var.get(), sheet_name_var.get(), starting_row_var.get(), ending_row_var.get()))
        elif function_box.get() == "Active to Completed":
            start_active_to_completed(load_excel(file_var.get(), sheet_name_var.get(), starting_row_var.get(), ending_row_var.get()))
        sleep(2)  # Wait for the operations to complete
        CTkMessagebox(title="Success", message="Automation completed successfully!", icon="check")
    except Exception as e:
        if isinstance(e, NameError):
            CTkMessagebox(title="Error", message="Please open ESMT first.", icon="cancel")
        else:
            CTkMessagebox(title="Error", message=f"An error occurred: {e}", icon="cancel")

# ------------------------ UI Layout ------------------------
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")
app = ctk.CTk()
app.title("ESMT - Automation")
app.geometry("650x750")

frame_1 = ctk.CTkFrame(app)
frame_2 = ctk.CTkFrame(app)
frame_3 = ctk.CTkFrame(app)

# ------------------------ Select Functionality Frame Starts ------------------------

ctk.CTkLabel(frame_1, text="What would you like to do?").grid(row=0, column=0, sticky="w", padx=(30,10), pady=(30,10))
function_box = ctk.CTkComboBox(frame_1, values=["Add Drawings", "Active to Completed", "Convert to Excel (WorkTime)"], command=select_functionality, width=250, height=30, border_color="#0a6488", button_color="#0a6488", border_width=2, text_color="#ffffff", dropdown_font=("Arial", 14), justify="center", dropdown_hover_color="#0a6488")
function_box.grid(row=0, column=1, sticky="w", padx=10, pady=(30,10))
function_box.set("Add Drawings")
select_functionality("Add Drawings")  # Default selection

# Show Instructions Button
ctk.CTkButton(frame_1, text="Show Instructions", command=show_instructions, fg_color="#188411", border_color="#d3eef7", border_width=2).grid(row=2, column=0, padx=(30,20), pady=(30,10))

# ------------------------ Select Functionality Frame Ends ------------------------


# ------------------------ Convert Excel Frame Starts ------------------------

# Select Excel File
ctk.CTkLabel(frame_3, text="Select Excel").grid(row=1, column=0, sticky="w", padx=(30,10), pady=10, ipady=20)
file_var2 = ctk.StringVar()
ctk.CTkEntry(frame_3, textvariable=file_var2, width=250).grid(row=1, column=1, sticky="w", padx=10, pady=10)
ctk.CTkButton(frame_3, text="Browse", command=lambda: browse_file(file_var2), width=60).grid(row=1, column=2, sticky="w", padx=10, pady=10)

# Convert Button
ctk.CTkButton(frame_3, text="Convert", command=lambda: convert_excel(file_var2.get()), fg_color="#0a6488", border_color="#d3eef7", border_width=2).grid(row=4, column=1, pady=20)

# ------------------------ Convert Excel Frame Ends -------------------------


# ------------------------ Shared Frame Starts ------------------------

# Select Excel File
ctk.CTkLabel(frame_2, text="Select Excel").grid(row=1, column=0, sticky="w", padx=(30,10), pady=10, ipady=10)
file_var = ctk.StringVar()
ctk.CTkEntry(frame_2, textvariable=file_var, width=250).grid(row=1, column=1, sticky="w", padx=10, pady=10)
ctk.CTkButton(frame_2, text="Browse", command=lambda: browse_file(file_var), width=60).grid(row=1, column=2, sticky="w", padx=10, pady=10)

# Sheet Name
sheet_name_var = ctk.StringVar()
ctk.CTkLabel(frame_2, text="Enter Sheet Name").grid(row=2, column=0, sticky="w", padx=(30,10), pady=10)
ctk.CTkEntry(frame_2, textvariable=sheet_name_var, width=150).grid(row=2, column=1, sticky="w", padx=10, pady=10)

# Starting Row Numbers
starting_row_var = ctk.StringVar()
ctk.CTkLabel(frame_2, text="Starting Row Number").grid(row=3, column=0, sticky="w", padx=(30,10), pady=10)
starting_row_var.set("3")  # Default value
vcmd = (frame_2.register(only_integers), '%P')
ctk.CTkEntry(frame_2, textvariable=starting_row_var, validate='key', validatecommand=vcmd, width=150).grid(row=3, column=1, sticky="w", padx=10, pady=10)

# Ending Row Numbers
ending_row_var = ctk.StringVar()
ctk.CTkLabel(frame_2, text="Ending Row Number").grid(row=4, column=0, sticky="w", padx=(30,10), pady=10)
ending_row_var.set("3")  # Default value
vcmd = (frame_2.register(only_integers), '%P')
ctk.CTkEntry(frame_2, textvariable=ending_row_var, validate='key', validatecommand=vcmd, width=150).grid(row=4, column=1, sticky="w", padx=10, pady=10)

#Chrome Button
ctk.CTkButton(frame_2, text="Launch ESMT", command=launch_chrome, border_color="#d3eef7", border_width=2).grid(row=5, column=0, padx=(30,20), pady=(30,10))

# Start Automation Button
ctk.CTkButton(frame_2, text="Start Automation", command=lambda: start_button_clicked(), fg_color="#0a6488", border_color="#d3eef7", border_width=2).grid(row=5, column=1, padx=(30,20), pady=(30,10))

# ------------------------ Shared Frame Ends ------------------------

app.mainloop()